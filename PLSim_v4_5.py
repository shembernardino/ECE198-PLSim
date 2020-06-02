#PLSim Interface
#row start is 1
#column start is 0

from tkinter import *
from tkinter import filedialog
import subprocess
import os
import io
from PIL import Image, ImageDraw, ImageDraw2
import cv2
import numpy
import numpy as np
import functools
#import pyexcel as pe
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, Color, PatternFill
from openpyxl.styles import colors 
import math
import matplotlib as mpl
import matplotlib.pyplot as plt

#CANVAS WIDTH = 800, HEIGHT = 300
APP_TITLE = "Drag & Drop Tk Canvas Images"
APP_XPOS = 500
APP_YPOS = 100
APP_WIDTH = 950
APP_HEIGHT = 750
#PIL DIMENSIONS
width = 800
height = 300
center = height//2
white = (255, 255, 255)
#green = (0,255,0)
blue = (135, 206, 235)
pink = (255,20,147)
purple = (144,238,144)

white = (255, 255, 255)
green = (0, 128, 0)           #Actual image color
brown = (42, 42, 165)         #Actual image color
orange = (100, 165, 255)
black = (0, 0, 0)
red = (0, 0, 255)
gray = (128, 128, 128)
antenna_height = [0, 179]

f = 868000000
speed_of_light = 300000000
max_cell_value = 0
min_cell_value = 100
#building_tip_flag = False
tip = None
building_width = []
building_height = []
building_angle = []
target_flag = []
line_slope = []
angle_arctan = []
tip_list = []
left_boundary_list = []
foliage_distance_list = []
first_terrain_coordinate = []

button_value = 0
scaling_value = 1
type_of_environment = 0
 
IMAGE_PATH = "C:/Users/Shem Bernardino/OneDrive/Documents/198Files/GUIFiles/"
 
##########################################################     INPUT FIELDS      ################################################################################################

def Input_Fields(window):   

    #def submit_values():
    #    freq = textentry1.get()
    #    ant_height = textentry2.get()
    #    max_distance = textentry3.get()
    #    #max_height = textentry4.get()
    #    pl_model = rx.get()

    #    print(freq, ant_height, max_distance, pl_model)

    #Input Parameters 
    In_Pam_x = 10
    In_Pam_y = 50
    Label1 = Label(window, text="Input Parameters", bg="black", fg="white", font="none 12 bold") 
    Label1.place(x=In_Pam_x, y=In_Pam_y)

    #Frequency
    Label2 = Label(window, text="Frequency:", bg="black", fg="white", font="none 10 bold") 
    Label2.place(x=In_Pam_x+20, y=In_Pam_y+30)

    #Frequency Entry Box
    textentry1 = Entry(window, width=20, bg="white")
    textentry1.place(x=In_Pam_x+120, y=In_Pam_y+30)

    #Antenna Height
    Label3 = Label(window, text="Antenna Height:", bg="black", fg="white", font="none 10 bold") 
    Label3.place(x=In_Pam_x+20, y=In_Pam_y+60)

    #Antenna Height Entry Box
    textentry2 = Entry(window, width=20, bg="white")
    textentry2.place(x=In_Pam_x+150, y=In_Pam_y+60)

    #Domain Parameters
    Dom_Pam_x = 350
    Dom_Pam_y = 50
    Label4 = Label(window, text="Domain Parameters", bg="black", fg="white", font="none 12 bold") 
    Label4.place(x=Dom_Pam_x, y=Dom_Pam_y)

    #Maximum Distance
    Label5 = Label(window, text="Scaling Factor:", bg="black", fg="white", font="none 10 bold") 
    Label5.place(x=Dom_Pam_x+20, y=Dom_Pam_y+30)

    #Maximum Distance Entry Box
    textentry3 = Entry(window, width=20, bg="white")
    textentry3.place(x=Dom_Pam_x+190, y=Dom_Pam_y+30)

    #Maximum Height
    #Label6 = Label(window, text="Maximum Height:", bg="black", fg="white", font="none 10 bold") 
    #Label6.place(x=Dom_Pam_x+20, y=Dom_Pam_y+60)

    #Maximum Height Entry Box
    #textentry4 = Entry(window, width=20, bg="white")
    #textentry4.place(x=Dom_Pam_x+170, y=Dom_Pam_y+60)

    ###########Radio Button########################

    #Path Loss Models
    Path_Mod_x = 750
    Path_Mod_y = 50
    Label7 = Label(window, text="Path Loss Models", bg="black", fg="white", font="none 12 bold") 
    Label7.place(x=Path_Mod_x, y=Path_Mod_y)

    rx = IntVar()

    #Radio Button Choices
    radio1 = Radiobutton(window, text="COST231 WI + MED", variable=rx, value=0) 
    radio1.place(x=Path_Mod_x+20, y=Path_Mod_y+30)
    radio2 = Radiobutton(window, text="COST231 WI + Tewari", variable=rx, value=1)
    radio2.place(x=Path_Mod_x+20, y=Path_Mod_y+60)
    #radio3 = Radiobutton(window, text="WINNER+", variable=rx, value=3) 
    #radio3.place(x=Path_Mod_x+20, y=Path_Mod_y+90)
    #radio4 = Radiobutton(window, text="Default", variable=rx, value=4) 
    #radio4.place(x=Path_Mod_x+20, y=Path_Mod_y+120)

    cx = IntVar()

    #Type of Environement
    radio3 = Radiobutton(window, text="Urban/Suburban", variable=cx, value=0) 
    radio3.place(x=Dom_Pam_x+20, y=Path_Mod_y+70)
    radio4 = Radiobutton(window, text="Dense Urban", variable=cx, value=1) 
    radio4.place(x=Dom_Pam_x+20, y=Path_Mod_y+100)

    #submit_button = Button(window, text="SUBMIT", width=6, command=submit_values) 
    #submit_button.place(x=50, y=120)

    ##################Signal Obstructions###################

    #Signal Obstructions
    Sig_Obs_x = 10
    Sig_Obs_y = 200
    Label8 = Label(window, text="Signal Obstructions", bg="black", fg="white", font="none 12 bold") 
    Label8.place(x=Sig_Obs_x, y=Sig_Obs_y)

    return textentry1, textentry2, textentry3, rx, cx

####################################################### IMAGE ANALYSER ###########################################################################################

def Line_for_Signal_Path(image, image_):

    def line(x0, y0, x1, y1):
        "Bresenham's line algorithm"
        points_in_line = []
        dx = abs(x1 - x0)
        dy = abs(y1 - y0)
        x, y = x0, y0
        sx = -1 if x0 > x1 else 1
        sy = -1 if y0 > y1 else 1
        if dx > dy:
            err = dx / 2.0
            while x != x1:
                points_in_line.append((x, y))
                err -= dy
                if err < 0:
                    y += sy
                    err += dx
                x += sx
        else:
            err = dy / 2.0
            while y != y1:
                points_in_line.append((x, y))
                err -= dx
                if err < 0:
                    x += sx
                    err += dy
                y += sy
        points_in_line.append((x, y))
        return points_in_line
        #print(points_in_line)
    
    image_dim = image_.size
    image_width = image_dim[0] - 2
    image_height = image_dim[1] - 2
    #print(image_width, image_height)
    x_pixel = 0
    y_pixel = 0
    counter = 0
    all_color_list = []
    all_points_list = []
    pts_colors = []

    pts = line(antenna_height[0], antenna_height[1], x_pixel, y_pixel)                                    #(0, 150) is the antenna coordinates
  
    #Run from x = 0 -> image width
    while x_pixel < image_width:
        all_points_list.append(pts)
        while counter < len(pts):
            image_coordinates = convert_to_image_coordinates(pts, counter)                #Convert (x,y) coordinates to (y,x), which is the coordinate system for the image itself
            pts_colors.append(image[image_coordinates[0],image_coordinates[1]])
            counter +=1

        counter = 0
        #Determine_Path_Loss(image, pts_colors, pts, ws1)
        all_color_list.append(pts_colors)
        pts_colors = []
        x_pixel +=1
        pts = line(antenna_height[0], antenna_height[1], x_pixel, y_pixel)

    y_pixel = 0
    counter = 0
    pts_colors = []
    pts = line(antenna_height[0], antenna_height[1], image_width-1, y_pixel)                                    #(0, 150) is the antenna coordinates
  
    #Run from y = 0 -> image height
    while y_pixel < image_height:
        all_points_list.append(pts)
        while counter < len(pts):
            image_coordinates = convert_to_image_coordinates(pts, counter)                #Convert (x,y) coordinates to (y,x), which is the coordinate system for the image itself
            pts_colors.append(image[image_coordinates[0],image_coordinates[1]])
            counter +=1

        counter = 0
        #Determine_Path_Loss(image, pts_colors, pts, ws1)
        all_color_list.append(pts_colors)
        pts_colors = []
        y_pixel +=1
        pts = line(antenna_height[0], antenna_height[1], image_width-1, y_pixel)

    x_pixel = image_width - 1
    y_pixel = image_height - 1
    counter = 0
    pts_colors = []
    pts = line(antenna_height[0], antenna_height[1], x_pixel, y_pixel)                                    #(0, 150) is the antenna coordinates
  
    #Run from x from image width -> 0
    while x_pixel >= 0:
        all_points_list.append(pts)
        while counter < len(pts):
            image_coordinates = convert_to_image_coordinates(pts, counter)                #Convert (x,y) coordinates to (y,x), which is the coordinate system for the image itself
            pts_colors.append(image[image_coordinates[0],image_coordinates[1]])
            counter +=1

        counter = 0
        #Determine_Path_Loss(image, pts_colors, pts, ws1)
        all_color_list.append(pts_colors)
        pts_colors = []
        x_pixel = x_pixel - 1
        pts = line(antenna_height[0], antenna_height[1], x_pixel, y_pixel)

    return all_points_list, all_color_list

########################################################## CONVERT TO IMAGE COORDINATES #########################################################################

def convert_to_image_coordinates(list, int):
    original_coordinates = list[int]

    xcoord = functools.reduce(lambda x: x*1, original_coordinates[0:1])
    ycoord = functools.reduce(lambda x: x*1, original_coordinates[1:2])

    x_image = ycoord
    y_image = xcoord

    return x_image, y_image

########################################################## DETERMINE PATH LOSS ##################################################################################

def Determine_Path_Loss(image, list, list_, sheet, list_2, list_3):

    master_points_list = list_2
    master_colors_list = list_3
    global building_height
    global tip
    global target_flag
    global building_angle

    def Modified_Exponential_Decay_Model(int1):
        foliage_distance = int1*scaling_value

        if foliage_distance <= 14:
            L = 0.45*((f/1000000)**0.284)*foliage_distance
        
        elif foliage_distance > 14:
            L = 1.33*((f/1000000)**0.284)*(foliage_distance**0.588)
        
        return L

    def Tewari_Model(float1):
        
        width_of_trees = (float1 + 1)*scaling_value
        first_term = -27.56
        second_term = 20*math.log10(f/1000000)

        #Alpha is negative to account for the negative sign in the exponential
        A_horizontal = 0.4491              
        A_vertical = 0.2661
        B_horizontal = 0.6291
        B_vertical = 0.5331
        alpha_horizontal = -0.0152
        alpha_vertical = 0.0140

        exponential = math.exp(alpha_horizontal*width_of_trees)
        log_arg = ((A_horizontal*exponential)/width_of_trees)+(B_horizontal/(width_of_trees*width_of_trees))    
        third_term = -20*math.log10(log_arg)

        tewari_result = first_term + second_term + third_term

        return tewari_result

    def find_foliage_width(image, coordinates):
        #Coordinates are still in (x,y) form at this point    
        orig_x = coordinates[0]
        orig_y = coordinates[1]
        x = orig_x
        pixel_color = image[orig_y, orig_x]

        #from current pixel to the right
        while np.all(pixel_color == green):
            x +=1
            pixel_color = image[orig_y, x]

            if np.all(pixel_color != green):
                x -=1
                break

        right_boundary = x
        x = orig_x

        #from current pixel to the left
        while np.all(pixel_color == green):
            x -=1
            pixel_color = image[orig_y, x]

            if np.all(pixel_color != green):
                x +=1
                break

        left_boundary = x

        distance = right_boundary - left_boundary
    
        return distance
    
    def COST_231_Walfisch_Ikegami_Model(tuple1, int1, int2, tuple2, int3, int4):        #Model does not include free space path loss yet; Add fspl in main code
        
        tip_coordinate = tuple1
        height_of_building = int1
        angle_of_signal_arrival = int2
        current_coordinates = tuple2
        distance_between_buildings = int3
        street_width = int4
        #no_more_next_building = flag

        Hb = height_of_building*scaling_value
        hm = ((tip_coordinate[1] + (Hb/scaling_value)) - current_coordinates[1])*scaling_value      #Y-component of tip plus building height equals y-component of base of building; Y-component of current coordinate minus y-component of base equals true height
        distance = (current_coordinates[0]*scaling_value)/1000
        hb = (first_terrain_coordinate[1] - antenna_height[1])*scaling_value        

        if type_of_environment == 0:             
            k_f = -4 + 0.7*(((f/1000000)/925)-1)
        elif type_of_environment == 1: 
            k_f = -4 + 1.5*(((f/1000000)/925)-1)

        if hb > Hb:
            L_bsh = -18*(math.log10(hb-Hb+1))
            k_a = 54
            k_d = 18

        elif hb <= Hb:
            L_bsh = 0
            k_d = 18 - 15*((hb-Hb)/Hb)

            if distance >= 0.5:
                k_a = 54 - 0.8*(hb-Hb)
            elif distance < 0.5:
                k_a = 54 - 1.6*(hb-Hb)*distance
        
        #if no_more_next_building == True:
        #    L_msd = L_bsh + k_a + k_d*(math.log10(distance)) + k_f*(math.log10(f/1000000))
        #else:
        L_msd = L_bsh + k_a + k_d*(math.log10(distance)) + k_f*(math.log10(f/1000000)) - 9*(math.log10(distance_between_buildings*scaling_value))

        delta_height = Hb - hm
        
        if delta_height <= 1:
            delta_height = 1

        L_ori = 0

        if angle_of_signal_arrival > 0 and angle_of_signal_arrival < 35:
            L_ori = -10 + 0.354*(angle_of_signal_arrival)

        elif angle_of_signal_arrival >= 35 and angle_of_signal_arrival < 55:
            L_ori = 2.5 + 0.075*(angle_of_signal_arrival - 35)

        elif angle_of_signal_arrival >= 55 and angle_of_signal_arrival <= 90:
            L_ori = 4 - 0.114*(angle_of_signal_arrival - 55)

        #if no_more_next_building == True:
        #    L_rts = -16.9 + 10*(math.log10(f/1000000)) + 20*(math.log10(delta_height)) + L_ori
        #else: 
        L_rts = -16.9 - 10*(math.log10(street_width*scaling_value)) + 10*(math.log10(f/1000000)) + 20*(math.log10(delta_height)) + L_ori

        Total_path_loss = L_msd + L_rts

        return Total_path_loss
        

    def find_building_height(image, gray_pixel_coordinate):
        #Coordinates are still in (x,y) form at this point    
        orig_x = gray_pixel_coordinate[0]
        orig_y = gray_pixel_coordinate[1]
        y = orig_y
        pixel_color = image[orig_y, orig_x]

        #from current pixel to the bottom
        while np.all(pixel_color != black) or np.all(pixel_color != white):
            y +=1
            pixel_color = image[y, orig_x]

            if np.all(pixel_color == black) or np.all(pixel_color == white):
                y -=1
                break

        bottom_boundary = y
        y = orig_y

        #from current pixel to the top
        while np.all(pixel_color != black) or np.all(pixel_color != white):
            y -=1
            pixel_color = image[y, orig_x]

            if np.all(pixel_color == black) or np.all(pixel_color == white):
                y +=1
                break

        top_boundary = y

        height = bottom_boundary - top_boundary
        return height

            

    def find_building_tip(image, coordinates):

        #print("Starting point is:", coordinates)
        x = coordinates[0]
        y = coordinates[1]
        pixel_color = image[y,x]

        while np.all(pixel_color == gray):
            #print("Moving right")
            x +=1
            pixel_color = image[y,x]

            if np.all(pixel_color != gray):
                x = x - 1
                break

        while np.all(pixel_color == gray):
            #print("Moving up")
            y -=1
            pixel_color = image[y,x]

            if np.all(pixel_color != gray):
                y = y - 1
                break

        pixel_color = image[y, x+1]

        while np.all(pixel_color == gray):
            #print("Moving right")
            x +=1
            pixel_color = image[y,x]

            if np.all(pixel_color != gray):
                x = x - 1
                break

        #print("Building tip coordinate is:", x, y)
        return (y, x) #Convert (y,x) back to (x,y)

    def find_second_coordinate(coordinate_list, counter):

        first_coordinate = coordinate_list[counter]
        x1 = first_coordinate[0]
        y1 = first_coordinate[1]

        c = 1
        second_coordinate = coordinate_list[counter-c]
        x2 = second_coordinate[0]
        y2 = second_coordinate[1]

        delta_x = x2 - x1
        delta_y = y2 - y1

        while delta_x == 0 or delta_y == 0:
            c +=1
            second_coordinate = coordinate_list[counter-c]
            x2 = second_coordinate[0]
            y2 = second_coordinate[1]

            delta_x = x2 - x1
            delta_y = y2 - y1

        return second_coordinate

    def find_angle(first_coordinate, second_coordinate):

        x2 = first_coordinate[0]
        y2 = first_coordinate[1]
        x1 = second_coordinate[0]
        y1 = second_coordinate[1]

        #print("First coordinate is:", first_coordinate)
        #print("Second coordinate is:", second_coordinate)

        delta_y = y2-y1
        delta_x = x2-x1

        slope = delta_y/delta_x
        global line_slope
        line_slope_ = slope

        if len(tip_list) != len(line_slope):
            if len(line_slope) == 0:
                line_slope.append(line_slope_)
            elif len(line_slope) < len(tip_list):
                line_slope.append(line_slope_)
                        
        elif len(tip_list) == len(line_slope):
            pass
        #print("Slope of line is:", slope)

        pi = math.pi
        arctan = math.atan(slope)
        angle2 = arctan*(180/pi)
        global angle_arctan
        angle_arctan_ = arctan

        if len(tip_list) != len(angle_arctan):
            if len(angle_arctan) == 0:
                angle_arctan.append(angle_arctan_)
            elif len(angle_arctan) < len(tip_list):
                angle_arctan.append(angle_arctan_)
                        
        elif len(tip_list) == len(angle_arctan):
            pass
        #print("Arctan of slope is:", arctan)

        return abs(angle2)

    def find_building_width(image, coordinates):
        #Coordinates are still in (x,y) form at this point    
        orig_x = coordinates[0]
        orig_y = coordinates[1]
        x = orig_x
        pixel_color = image[orig_y, orig_x]

        #from current pixel to the right
        while np.all(pixel_color == gray):
            x +=1
            #print(x)
            pixel_color = image[orig_y, x]

            if np.all(pixel_color != gray):
                x -=1
                break

        right_boundary = x
        x = orig_x

        #from current pixel to the left
        while np.all(pixel_color == gray):
            x -=1
            pixel_color = image[orig_y, x]

            if np.all(pixel_color != gray):
                x +=1
                break

        left_boundary = x

        width = right_boundary - left_boundary
        width_over_two = width/2
        building_center_coordinate = left_boundary + width_over_two
    
        return building_center_coordinate, left_boundary



    color_values = []
    color_values = list                 #List for the color values of the image per pixel in the line
    pixel_coordinates = []
    pixel_coordinates = list_           #List for the pixel coordinates in the line
    tree_distance = 0
    green_flag = False
    terrain_flag = False
    gray_flag = False
    global tip_list
    global building_height 
    global building_angle 
    global target_flag 
    global line_slope 
    global angle_arctan 
    k = 0
    COST_231_model_list = []
    MED_values = []
    TMV_values = []
    copy_flag = False
    copy_flag_2 = False

    #wb = Workbook()
    #ws1 = wb.create_sheet("Path Loss Values")

    counter = 0                         #Counter scale = 1 = 1m
    global building_width 
    global left_boundary_list
    building_width = []
    left_boundary_list = []
    global foliage_distance_list
    foliage_distance_list = []

    while counter < len(list):
        if np.all(color_values[counter] == gray):
            if np.all(color_values[counter-1] != gray)  and counter != 0:
                center_coordinate, left_bound = find_building_width(image, pixel_coordinates[counter])
                if len(building_width) == 0:
                    building_width.append(center_coordinate)
                    left_boundary_list.append(left_bound)
                elif len(building_width) > 0:
                    t = 0
                    current_width = building_width[t]
                    if center_coordinate != current_width:
                        while t < len(building_width):
                            if center_coordinate == building_width[t]:
                                copy_flag_2 = True
                                break
                            else:
                                t +=1

                        if copy_flag_2 == False:
                            building_width.append(center_coordinate)
                            left_boundary_list.append(left_bound)
                        else:
                            pass
                        
        copy_flag_2 = False
        counter +=1

    #if len(building_width) > 0:
    #    print("Building width is:", building_width)
    #    print("Left boundary list is:", left_boundary_list)

    counter = 0

    while counter < len(list):
        
        cell_coordinates = convert_to_image_coordinates(pixel_coordinates, counter)                     #Convert (x,y) (the coordinates of the line) to (y,x)
        #print(cell_coordinates)
        current_cell = sheet.cell(row=cell_coordinates[0]+1, column=cell_coordinates[1]+1)
        xy_coordinates = pixel_coordinates[counter]
        global min_cell_value
        global max_cell_value
        real_distance = counter*scaling_value

        if np.all(color_values[counter] == white):  
            current_cell.fill = PatternFill(fill_type='solid', start_color='00FFFF', end_color='00FFFF')

            if counter == 0:
                current_cell.value = 0

            elif green_flag == True and gray_flag == False:
                MED_values.clear()
                TMV_values.clear()

                if button_value == 0:
                    #u = 0
                    #while u < len(foliage_distance_list):
                    #    tree_distance = foliage_distance_list[u]
                    #    Modified_Exponential_Decay = Modified_Exponential_Decay_Model(tree_distance)
                    #    MED_values.append(Modified_Exponential_Decay)
                    #    u +=1

                    #Foliage_Path_Loss = sum(MED_values)

                    Modified_Exponential_Decay = Modified_Exponential_Decay_Model(tree_distance)
                    Foliage_Path_Loss = Modified_Exponential_Decay

                elif button_value == 1:
                    #u = 0
                    #while u < len(foliage_distance_list):
                    #    tree_distance = foliage_distance_list[u]
                    #    Tewari_Model_Value = Tewari_Model(tree_distance)
                    #    TMV_values.append(Tewari_Model_Value)
                    #    u +=1

                    #Foliage_Path_Loss = sum(TMV_values)
                    Tewari_Model_Value = Tewari_Model(tree_distance)
                    Foliage_Path_Loss = Tewari_Model_Value
                    
                Free_space_path_loss = 32.44 + 20*(math.log10(real_distance/1000))  + 20*(math.log10(f/1000000))
                current_cell.value = Foliage_Path_Loss + Free_space_path_loss

                if current_cell.value < min_cell_value:
                    if  current_cell.value > 3:
                        min_cell_value = current_cell.value

                elif current_cell.value > max_cell_value:
                    max_cell_value = current_cell.value
            
            elif gray_flag == True and green_flag == False:
                a = 0
                COST_231_model_list.clear()
                while a < len(tip_list):
                    current_tip = tip_list[a]
                    if xy_coordinates[0] > current_tip[0] and xy_coordinates[1] >= current_tip[1]:
                        p = 0
                        less_than_list = []
                        more_than_list = []

                        while p < len(building_width):                  #Find the two nearest building centers from the coordinate
                            if current_tip[0] < building_width[p]:
                                less_than_list.append(building_width[p])
                            elif current_tip[0] > building_width[p]:
                                more_than_list.append(building_width[p])
                            p +=1

                        #if len(less_than_list) == 0 or len(more_than_list) == 0:
                        if len(less_than_list) == 0:
                            less_than_list.append(800)

                        right_building_center = min(less_than_list)

                        if len(more_than_list) == 0:
                            more_than_list.append(0)
                
                        left_building_center = max(more_than_list)
                        #print(less_than_list, right_building_center)
                        #print(more_than_list, left_building_center)

                        building_center_distance = right_building_center - left_building_center

                        p = 0
                        less_than_list = []

                        while p < len(left_boundary_list):                  #Find the street width
                            if current_tip[0] < left_boundary_list[p]:
                                less_than_list.append(left_boundary_list[p])
                            p +=1

                        if len(less_than_list) == 0:
                            less_than_list.append(800)

                        right_building_edge = min(less_than_list)

                        street_width = right_building_edge - current_tip[0]

                        #if xy_coordinates[0] == 250:
                        #    print("Current point is:", xy_coordinates)
                        #    print("Tip list is:", tip_list)
                        #    print("Current tip is:", current_tip)
                        #    print("Building width is:", building_width)
                        #    print("Left boundary list is:", left_boundary_list)
                        #    print("Right Building Edge is:", right_building_edge)
                        #    print("Right Building Center is:", right_building_center)
                        #    print("Left Building Center is:", left_building_center)
                        #    print(building_center_distance, street_width)
                        #    print(" ")

                        if right_building_center == 800:
                            building_center_distance = 7
        
                        if right_building_edge == 800:
                            street_width = 7

                        COST_231_model = COST_231_Walfisch_Ikegami_Model(tip_list[a], building_height[a], building_angle[a], pixel_coordinates[counter], building_center_distance, street_width)
                        COST_231_model_list.append(COST_231_model)
                        a +=1
                    else: 
                        a +=1

                #print(COST_231_model_list)
                COST_231_model_total = sum(COST_231_model_list)
                Free_space_path_loss = 32.44 + 20*(math.log10(real_distance/1000))  + 20*(math.log10(f/1000000))
                current_cell.value = COST_231_model_total + Free_space_path_loss

                if current_cell.value < min_cell_value:
                    if  current_cell.value > 3:
                        min_cell_value = current_cell.value

                elif current_cell.value > max_cell_value:
                    max_cell_value = current_cell.value

            elif gray_flag == True and green_flag == True:
                a = 0
                COST_231_model_list.clear()
                while a < len(tip_list):
                    current_tip = tip_list[a]
                    if xy_coordinates[0] > current_tip[0] and xy_coordinates[1] >= current_tip[1]:
                        p = 0
                        less_than_list = []
                        more_than_list = []

                        while p < len(building_width):                  #Find the two nearest building centers from the coordinate
                            if current_tip[0] < building_width[p]:
                                less_than_list.append(building_width[p])
                            elif current_tip[0] > building_width[p]:
                                more_than_list.append(building_width[p])
                            p +=1

                        #if len(less_than_list) == 0 or len(more_than_list) == 0:
                        if len(less_than_list) == 0:
                            less_than_list.append(800)

                        right_building_center = min(less_than_list)

                        if len(more_than_list) == 0:
                            more_than_list.append(0)
                
                        left_building_center = max(more_than_list)
                        #print(less_than_list, right_building_center)
                        #print(more_than_list, left_building_center)

                        building_center_distance = right_building_center - left_building_center

                        p = 0
                        less_than_list = []

                        while p < len(left_boundary_list):                  #Find the street width
                            if current_tip[0] < left_boundary_list[p]:
                                less_than_list.append(left_boundary_list[p])
                            p +=1

                        if len(less_than_list) == 0:
                            less_than_list.append(800)

                        right_building_edge = min(less_than_list)

                        street_width = right_building_edge - current_tip[0]

                        #if xy_coordinates[0] == 250:
                        #    print("Current point is:", xy_coordinates)
                        #    print("Tip list is:", tip_list)
                        #    print("Current tip is:", current_tip)
                        #    print("Building width is:", building_width)
                        #    print("Left boundary list is:", left_boundary_list)
                        #    print("Right Building Edge is:", right_building_edge)
                        #    print("Right Building Center is:", right_building_center)
                        #    print("Left Building Center is:", left_building_center)
                        #    print(building_center_distance, street_width)
                        #    print(" ")

                        if right_building_center == 800:
                            building_center_distance = 7
        
                        if right_building_edge == 800:
                            street_width = 7

                        COST_231_model = COST_231_Walfisch_Ikegami_Model(tip_list[a], building_height[a], building_angle[a], pixel_coordinates[counter], building_center_distance, street_width)
                        COST_231_model_list.append(COST_231_model)
                        a +=1
                    else: 
                        a +=1

                #print(COST_231_model_list)
                COST_231_model_total = sum(COST_231_model_list)

                MED_values.clear()
                TMV_values.clear()

                if button_value == 0:
                    #u = 0
                    #while u < len(foliage_distance_list):
                    #    tree_distance = foliage_distance_list[u]
                    #    Modified_Exponential_Decay = Modified_Exponential_Decay_Model(tree_distance)
                    #    MED_values.append(Modified_Exponential_Decay)
                    #    u +=1

                    #Foliage_Path_Loss = sum(MED_values)

                    Modified_Exponential_Decay = Modified_Exponential_Decay_Model(tree_distance)
                    Foliage_Path_Loss = Modified_Exponential_Decay

                elif button_value == 1:
                    #u = 0
                    #while u < len(foliage_distance_list):
                    #    tree_distance = foliage_distance_list[u]
                    #    Tewari_Model_Value = Tewari_Model(tree_distance)
                    #    TMV_values.append(Tewari_Model_Value)
                    #    u +=1

                    #Foliage_Path_Loss = sum(TMV_values)
                    Tewari_Model_Value = Tewari_Model(tree_distance)
                    Foliage_Path_Loss = Tewari_Model_Value

                Free_space_path_loss = 32.44 + 20*(math.log10(real_distance/1000))  + 20*(math.log10(f/1000000))
                current_cell.value = COST_231_model_total + Foliage_Path_Loss + Free_space_path_loss

                if current_cell.value < min_cell_value:
                    if  current_cell.value > 3:
                        min_cell_value = current_cell.value

                elif current_cell.value > max_cell_value:
                    max_cell_value = current_cell.value

            else:
                #current_cell.value = 20*(math.log10((4*1*counter*f)/speed_of_light)) 
                Free_space_path_loss = 32.44 + 20*(math.log10(real_distance/1000))  + 20*(math.log10(f/1000000))
                current_cell.value = Free_space_path_loss

                if current_cell.value < min_cell_value:
                    if  current_cell.value > 3:
                        min_cell_value = current_cell.value

                elif current_cell.value > max_cell_value:
                    max_cell_value = current_cell.value

            counter +=1

        elif np.all(color_values[counter] == green):
            green_flag = True
            current_cell.value = 1
            current_cell.fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
            tree_distance +=1

            #if np.all(color_values[counter-1] != green) and counter != 0:
                #foliage_width = find_foliage_width(image, pixel_coordinates[counter])
                ##print("Foliage is:", foliage_width)
                #foliage_distance_list.append(foliage_width)

                #if len(foliage_distance_list) > 0:
                    #print("Foliage width is:", foliage_distance_list)

            counter +=1

        elif np.all(color_values[counter] == gray):

            gray_flag = True
            current_cell.value = 3
            current_cell.fill = PatternFill(fill_type='solid', start_color='808080', end_color='808080')

            if np.all(color_values[counter-1] != gray) and counter != 0:
                target_list = 0
                building_tip = find_building_tip(image, pixel_coordinates[counter])
                xcoord = functools.reduce(lambda x: x*1, building_tip[0:1])
                ycoord = functools.reduce(lambda x: x*1, building_tip[1:2])
                building_tip = []
                building_tip.append(ycoord)
                building_tip.append(xcoord)
                tip = tuple(building_tip)

                if len(tip_list) > 0:               #Need to change tip list algorithm
                    current_tip = tip_list[k]
                    #building_height_ = find_building_height(image, pixel_coordinates[counter])
                    if tip[0] != current_tip[0]:
                        j = 0
                        while j < len(tip_list):
                            tip_current = tip_list[j]
                            if tip[0] == tip_current[0]:
                                copy_flag = True
                                break
                            else:
                                j +=1
                        
                        if copy_flag == False:
                            tip_list.append(tip)
                            k +=1
                        else:
                            pass
                else:
                    tip_list.append(tip)
                
                #copy_flag = False

                building_height_ = find_building_height(image, pixel_coordinates[counter])

                if len(tip_list) != len(building_height):
                    if len(building_height) == 0:
                        building_height.append(building_height_)
                    elif len(building_height) < len(tip_list):
                        building_height.append(building_height_)

                elif len(tip_list) == len(building_height):
                    pass

                first_coordinate = tip
                #second_coordinate = find_second_coordinate(current_list, dbuff)
                second_coordinate = (antenna_height[0], antenna_height[1])
                building_angle_ = find_angle(first_coordinate, second_coordinate)

                if len(tip_list) != len(building_angle):
                    if len(building_angle) == 0:
                        building_angle.append(building_angle_)
                    elif len(building_angle) < len(tip_list):
                        building_angle.append(building_angle_)
                        
                elif len(tip_list) == len(building_angle):
                    pass

            else:
                pass
            
            copy_flag = False
            counter +=1

        elif np.all(color_values[counter] == black):
            terrain_flag = True
            current_cell.value = 2
            current_cell.fill = PatternFill(fill_type='solid', start_color='000000', end_color='000000')
            counter +=1

        else:
            counter +=1
            #print("Pixel is white!")
    #print(foliage_distance_list)
    counter = 0
    #print("Tree distance is:", tree_distance)

    #wb.save("Image Path Loss Values.xlsx")

#################################################################### ADD GRADIENT TO WORKING IMAGE ################################################################

def Add_Gradient_to_Image(image, sheet):

    #wb = load_workbook("Image Path Loss Values.xlsx")
    #sheet = wb.active
    x_counter = 1
    y_counter = 1
    new_image = image.copy()

    while x_counter < 300:
        while y_counter < 800:
            current_cell = sheet.cell(row=x_counter, column=y_counter)
            x_image = x_counter - 1
            y_image = y_counter - 1
            #print(current_cell.value)
            if current_cell.value == None:
                y_counter +=1
                continue

            elif current_cell.value == 1:
                #print("Cell is 1")
                y_counter +=1
                continue

            elif current_cell.value == 2:
                y_counter +=1
                continue

            elif current_cell.value == 3:
                y_counter +=1
                continue
            #global min_cell_value
            #global max_cell_value

            #current_cell.fill = PatternFill(fill_type='solid', start_color='FFA500', end_color='FFA500')

            c1 = [0, 255, 255] #yellow
            c2 = [22, 0, 147] #red
            pixel_color_value = [0, 0, 0]
            color_scale = c2[2] + c1[1]
            min_max_diff = max_cell_value - min_cell_value

            color_step = color_scale/min_max_diff           #Step for the color spectrum
            pixel_color_offset = (current_cell.value - min_cell_value)
            mix = pixel_color_offset/min_max_diff

            value1 = [(1-mix)*x for x in c1]
            value2 = [mix*i for i in c2]
            #print(value1)
            #print(value2)

            for i in range(0,3):
                pixel_color_value[i] = value1[i] + value2[i] 
            #print(pixel_color_value)

            new_image[x_image, y_image] = pixel_color_value

            y_counter +=1

        x_counter += 1
        y_counter = 1
        

    return new_image


#################################################################### CANVAS OPERATIONS ###########################################################################
def Canvas_Operations(window, canvas, list1):

    def load_file_terrain(window, canvas):
        terrain = load_terrain_coords(window, canvas)
        return terrain

    def load_tree():
        trees = load_tree_coords(window, canvas)

    def load_building():
        buildings = load_building_coords(window, canvas)

    def save_image():

        a = list1[0].get()
        if a == "":
            print("Frequency is set to default")
        elif a != "":
            global f
            f = int(list1[0].get())

        print("Initial terrain Coordinate is:", first_terrain_coordinate)

        a = list1[2].get()
        if a == "":
            print("Scaling value is set to default")
        elif a != "":
            global scaling_value
            scaling_value = int(list1[2].get())

        a = list1[1].get()
        if a == "":
            print("Antenna height is set to default")
        elif a != "":
            global antenna_height
            antenna_height[1] = first_terrain_coordinate[1] - int((int(list1[1].get())/scaling_value))  

        a = list1[3].get()
        if a == "":
            print("Mode is set to default")
        elif a != "":
            global button_value
            button_value = int(list1[3].get())
        
        a = list1[4].get()
        if a == "":
            print("Type of environment is set to default")
        elif a != "":
            global type_of_environment
            button_value = int(list1[4].get())

        print("Frequency is:", f)
        print("Antenna height is:", antenna_height)
        print("Scaling value is:", scaling_value)

        if button_value == 0:
            print("Mode is: COST231 WI + MED")
        elif button_value == 1:
            print("Mode is: COST231 WI + Tewari Model")

        if type_of_environment == 0: 
            print("Type of environment is: Urban/Suburban")
        elif type_of_environment == 1:
            print("Type of environment is: Dense Urban")
        
        ps = canvas.postscript(file="trial.eps", colormode='color')
        img = Image.open("trial.eps")           
        #img = Image.open(io.BytesIO(ps.encode('utf-8')))
        #img.resize((1000,500))
        img.save("Original_Image.png", "png")

        basewidth = 800                                 #Resize the image
        img = Image.open('Original_Image.png')
        wpercent = (basewidth/float(img.size[0]))
        hsize = int((float(img.size[1])*float(wpercent)))
        img = img.resize((basewidth,hsize), Image.NEAREST)
        img.save('Resized_Original_Image.png') 

        #pix_val = list(img.getdata())               #Get value of every pixel
        ##numpy.savetxt("new_image.csv", pix_val, delimiter=",")             #Save pixel values to csv file
        #f = open("image_matrix.txt","w")
        #f.write(str(pix_val))
        #f.close
        
        #img.show()
        working_image = cv2.imread("Resized_Original_Image.png", 1)
        working_image = np.delete(working_image, 0, 1)                      #Trim white pixels in the left border; 0 is the element number, 1 is the axis number (0 for row and 1 for column)
        working_image = np.delete(working_image, -1, 1)             #Trim the white pixels in the right border
        #print(working_image.size)
        cv2.imshow("Original Image", working_image) #show image #IMAGE SIZE IS 602x226 without resizing

        list_of_points, list_of_colors_of_points = Line_for_Signal_Path(working_image, img)     #Process all lines and save all points and colors in their respective lists

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Path Loss Values"
        
        #print("Length of points list is:", len(list_of_points))
        #print("Length of colors list is:", len(list_of_colors_of_points))

        counter = 0
        while counter < len(list_of_points):
            points = list_of_points[counter]
            colors = list_of_colors_of_points[counter]
            Determine_Path_Loss(working_image, colors, points, ws1, list_of_points, list_of_colors_of_points)          #Detemine all path loss values of each pixel and write it in an excel sheet
            counter +=1

        print("Building Height is:", building_height)
        print("Building Angle is:", building_angle)
        print("The tip coordinate is:", tip_list)
        print("The Slope is:", line_slope)
        print("The Arctan of the slope is:", angle_arctan)
        print("Is target acquisition successful?", target_flag)
        print("Max value is:", max_cell_value)
        print("Min value is:", min_cell_value)
        print("Tip list is:", tip_list)
        print("Building width is:", building_width)
        
        Edited_image = Add_Gradient_to_Image(working_image, ws1)     #Add color to image to reflect path loss values
        cv2.imshow("Path Loss Gradient", Edited_image)
        cv2.imwrite("Gradient_Path_Loss.png", Edited_image)

        if tip == None:
            print("Tip is None")
        else:
            print(working_image[tip[1], tip[0]])

        wb.save("Image Path Loss Values.xlsx")

        ########################## IMAGE MATRIX HANDLING #######################################
        #np.set_printoptions(threshold=np.inf)
        #numpy.savetxt("image.csv", scaled_image, delimiter=",")
        #f = open("image_matrix.txt","w")
        #f.write(str(scaled_image))
        #f.close

    class Terrain:
        def __init__(self, window, canvas):
            self.window = window
            self.canvas = canvas
            self.value = 0
        
            def load_terrain():
                self.value = load_file_terrain(window, canvas)
                return self.value

            load_terrain = Button(window, text="Load Terrain", width=15, command=load_terrain)    #Load button for loading coordinates for terrain profile
            load_terrain.place(x=650, y=310)

    A = Terrain(window, canvas)
    #A = Terrain.value
    
    load_trees = Button(window, text="Load Trees", width=15, command=load_tree)    #Load button for loading coordinates for terrain profile
    load_trees.place(x=650, y=280)

    load_buildings = Button(window, text="Load Buildings", width=15, command=load_building)    #Load button for loading coordinates for terrain profile
    load_buildings.place(x=650, y=250)

    load_drawing = Button(window, text="SIMULATE", width=15, command=save_image)
    load_drawing.place(x=780, y=250)

    #add a clear button
    clear_button = Button(window, text="CLEAR", width=6, command=canvas.delete(A)) 
    clear_button.place(x=70, y=330)


################################################# OBSTRUCTION OPTIONS #############################################################
def Obstruction_AddOns(window, canvas):

    obs_button_x = 100
    obs_button_y = 220
    
    s = IntVar()

    Label9 = Label(window, text="x1:", bg="black", fg="white", font="none 10 bold") 
    Label9.place(x=obs_button_x+150, y=obs_button_y+30)

    img_x1 = Entry(window, width=20, bg="white")
    img_x1.place(x=obs_button_x+180, y=obs_button_y+30)

    Label10 = Label(window, text="y1:", bg="black", fg="white", font="none 10 bold") 
    Label10.place(x=obs_button_x+150, y=obs_button_y+60)

    img_y1 = Entry(window, width=20, bg="white")
    img_y1.place(x=obs_button_x+180, y=obs_button_y+60)

    Label11 = Label(window, text="x2:", bg="black", fg="white", font="none 10 bold") 
    Label11.place(x=obs_button_x+330, y=obs_button_y+30)

    img_x2 = Entry(window, width=20, bg="white")
    img_x2.place(x=obs_button_x+360, y=obs_button_y+30)

    Label12 = Label(window, text="y2:", bg="black", fg="white", font="none 10 bold") 
    Label12.place(x=obs_button_x+330, y=obs_button_y+60)

    img_y2 = Entry(window, width=20, bg="white")
    img_y2.place(x=obs_button_x+360, y=obs_button_y+60)

    def add_item_building():
        x1 = img_x1.get()
        y1 = img_y1.get()
        x2 = img_x2.get()
        y2 = img_y2.get()

        if x1=='' or y1=='' or x2=='' or y2=='':
            x1 = 100
            y1 = 100
            x2 = 150
            y2 = 150
            image = CreateCanvasObject(canvas, x1, y1, x2, y2, 'gray')

        else:    
            image = CreateCanvasObject(canvas, x1, y1, x2, y2, 'gray')

            
    def add_item_tree():
        x1 = img_x1.get()
        y1 = img_y1.get()
        x2 = img_x2.get()
        y2 = img_y2.get()

        if x1=='' or y1=='' or x2=='' or y2=='':
            x1 = 100
            y1 = 100
            x2 = 150
            y2 = 150
            image = CreateCanvasObject(canvas, x1, y1, x2, y2, 'green')

        else:
            image = CreateCanvasObject(canvas, x1, y1, x2, y2, 'green')

    #def delete_image(canvas):
        #canvas.delete(object)

    #Button Choices
    tree_button = Button(window, text="Tree", width=6, command=add_item_tree) 
    tree_button.place(x=obs_button_x+20, y=obs_button_y+30)
    building_button = Button(window, text="Building", width=6, command=add_item_building) 
    building_button.place(x=obs_button_x+85, y=obs_button_y+30)

    #Delete Button
    #delete_button = Button(window, text="Undo", width=6, command=delete_image(canvas)) 
    #delete_button.place(x=obs_button_x+350, y=obs_button_y+30)

########################################################### LOAD BUILDINGS ##################################################################################

def load_building_coords(window, canvas):
    line_number = 0
    current_point = 0
    next_point = 1
    x = []
    y = []
    filename = filedialog.askopenfilename(initialdir="C:/Users/Shem Bernardino/OneDrive/Documents/198Files/SimulationFiles", title="Select a File", filetypes=(("txt files", "*.txt"),("all files", "*.*")))
    file_ = open(filename, "r")
    #file_ = open("building_coordinates.txt", "r")
    f = file_.readline()

    while f:
        line_number +=1

        if line_number % 2 == 0: 
            y.append(int(f))

        else:
            x.append(int(f))
                
        f = file_.readline()
        
    file_.close()

    #c=0
    #while c < len(x):
    #    print("Building %d is:" %c, x[c], y[c])
    #    c = c + 1
    building_object_list = []

    while next_point < len(x):
        building = CreateCanvasObject(canvas, x[current_point], y[current_point], x[next_point], y[next_point], 'gray')
        building_object_list.append(building)
        current_point +=2
        next_point +=2

    return building_object_list

######################################################################## LOAD TREES ########################################################################

def load_tree_coords(window, canvas):
    line_number = 0
    current_point = 0
    next_point = 1
    x = []
    y = []
    filename = filedialog.askopenfilename(initialdir="C:/Users/Shem Bernardino/OneDrive/Documents/198Files/SimulationFiles", title="Select a File", filetypes=(("txt files", "*.txt"),("all files", "*.*")))
    file_ = open(filename, "r")
    #file_ = open("tree_coordinates.txt", "r")
    f = file_.readline()

    while f:
        line_number +=1

        if line_number % 2 == 0: 
            y.append(int(f))

        else:
            x.append(int(f))
                
        f = file_.readline()
        
    file_.close()

    #c=0
    #while c < len(x):
    #    print("Tree %d is:" %c, x[c], y[c])
    #    c = c + 1
    tree_object_list = []

    while next_point < len(x):
        tree = CreateCanvasObject(canvas, x[current_point], y[current_point], x[next_point], y[next_point], 'green')
        tree_object_list.append(tree)
        current_point +=2
        next_point +=2

    return tree_object_list
    

######################################################################## LOAD TERRAIN ############################################################################

def load_terrain_coords(window, canvas):
    poly_coords = []
    filename = filedialog.askopenfilename(initialdir="C:/Users/Shem Bernardino/OneDrive/Documents/198Files/SimulationFiles", title="Select a File", filetypes=(("txt files", "*.txt"),("all files", "*.*")))
    file_ = open(filename, "r")
    #file_ = open("terrain_coordinates.txt", "r")
    f = file_.readline()
    a = 0

    while f:
        poly_coords.append(int(f))
        if a < 2:
            global first_terrain_coordinate
            first_terrain_coordinate.append(int(f))
        f = file_.readline()
        a +=1
    
    poly_coords.append(800)                         #Append point in the lower righthand-side corner of canvas
    poly_coords.append(300)
    poly_coords.append(0)                           #Append point in the lower righthand-side corner of canvas
    poly_coords.append(300)
    poly_coords.append(poly_coords[0])              #Append first point of polygon to close it
    poly_coords.append(poly_coords[1])
    file_.close()

    polygon = canvas.create_polygon(poly_coords)

    return polygon

########################################################## END LOAD TERRAIN #############################################################################################

class CreateCanvasObject(object):
    def __init__(self, canvas, xpos1, ypos1, xpos2, ypos2, color):
        self.canvas = canvas
        #self.image_name = image_name
        self.xpos1, self.ypos1 = xpos1, ypos1
        self.xpos2, self.ypos2 = xpos2, ypos2
        self.color = color
 
        #self.tk_image = PhotoImage(
        #    file="{}{}".format(IMAGE_PATH, image_name))
        #self.image_obj= canvas.create_image(
        #    xpos, ypos, image=self.tk_image)

        self.image_obj = canvas.create_rectangle(xpos1, ypos1, xpos2, ypos2, fill = color, outline="") 
         
        canvas.tag_bind(self.image_obj, '<Button1-Motion>', self.move)
        canvas.tag_bind(self.image_obj, '<ButtonRelease-1>', self.release)
        self.move_flag = False
         
    def move(self, event):
        if self.move_flag:
            new_xpos, new_ypos = event.x, event.y
             
            self.canvas.move(self.image_obj,
                new_xpos-self.mouse_xpos ,new_ypos-self.mouse_ypos)
             
            self.mouse_xpos = new_xpos
            self.mouse_ypos = new_ypos
            #print(new_xpos, new_ypos)
            
        else:
            self.move_flag = True
            self.canvas.tag_raise(self.image_obj)   #move object on top of other objects
            self.mouse_xpos = event.x
            self.mouse_ypos = event.y
 
    def release(self, event):
        self.move_flag = False
                     
#class Signal_Obstructions(Frame):
 
    #def __init__(self, master):
        #self.master = master
        #self.master.protocol("WM_DELETE_WINDOW", self.close)
        #Frame.__init__(self, master)
 
        #self.canvas = Canvas(self, width=800, height=300, bg='white',
        #    highlightthickness=0)
        #self.canvas.pack(fill="both", expand=True)
         
        #self.image_1 = CreateCanvasObject(self.canvas, "tree.gif", 200, 200)
        #self.image_2 = CreateCanvasObject(self.canvas, "building.gif", 300, 200)
             
    #def close(self):
        #print("Application-Shutdown")
        #self.master.destroy()
     
def main():
    #close window function
    def close_window():
	    window.destroy()
	    exit()

    def clear_canvas(canvas):
        current_canvas = canvas
        current_canvas.delete("all")

    #app_win = Tk()
    window = Tk()
    window.title("PLSim")
    window.configure(background="black")
    window.geometry("{}x{}".format(APP_WIDTH, APP_HEIGHT))   #fixed app window size
    #app_win.title(APP_TITLE)
    window.geometry("+{}+{}".format(APP_XPOS, APP_YPOS)) #position of window in desktop
    #app = Application(app_win).pack(fill='both', expand=True)

    user_inputs = Input_Fields(window)                                               #Input all the text fields and texts

    buffer_label1 = Label(window, text="                    ", bg="black")
    buffer_label1.grid(row=1, column=0, sticky=W)

    mapper = Canvas(width=800, height=300, bg='white', highlightthickness=0)    #Canvas for the mapper
    mapper.grid(row=2, column=1, sticky=W)                                      #Canvas dimensions for mapper
    Obstruction_AddOns(window, mapper)                           #Function for drag and drop of obstruction images

    buffer_label2 = Label(window, text="\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n", bg="black")     #Spacing for mapper canvas
    buffer_label2.grid(row=0, column=0, sticky=W)

    Canvas_Operations(window, mapper, user_inputs)

    #add an exit button
    exit_button = Button(window, text="EXIT", width=6, command=close_window) 
    exit_button.place(x=APP_WIDTH-100, y=APP_HEIGHT-50)
     
    #window loop
    window.mainloop()
  
  
if __name__ == '__main__':
    main()   


#################     DND TRIAL END     #################################
